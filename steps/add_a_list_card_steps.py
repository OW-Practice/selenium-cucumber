import os
from behave import *
from openpyxl.workbook import Workbook
from pages.add_a_list_card_page import AddAListCardPage
from faker import Faker
from pathlib import Path

fake = Faker()
list_names = []
total_card_names = []
total_colors = []
total_label_names = []


@when(u'Create list title "{input}" and validate list name')
def create_list_title_validate_list_name(context, input):
    context.list_card = AddAListCardPage(context.driver)
    value = int(input)
    for i in range(value):
        name = fake.name()
        list_names.append(name)
        context.list_card.enter_list_title(name)
        context.list_card.click_on_add_a_list_button()
        context.list_card.verify_added_list_name(name)


@when(u'Create card with list size "{list_size}" and card size "{card_size}" and enter card name')
def click_on_add_a_card_button_enter_card_name(context, list_size, card_size):
    context.list_card = AddAListCardPage(context.driver)
    lists_val = int(list_size)
    cards_val = int(card_size)
    for i in range(lists_val):
        name_of_list = list_names[i]
        context.list_card.click_on_add_a_card_button(name_of_list)
        card_names = []
        for j in range(cards_val):
            name_card = fake.name()
            card_names.append(name_card)
            context.list_card.enter_card_name(name_card)
            context.list_card.click_on_add_card_button()
        total_card_names.append(card_names)


@then(u'Validate cards based on list size "{list_size}" and card size "{card_size}"')
def validate_created_cards_under_lists(context, list_size, card_size):
    context.list_card = AddAListCardPage(context.driver)
    lists_val = int(list_size)
    cards_val = int(card_size)
    for i in range(lists_val):
        name_of_list = list_names[i]
        for j in range(cards_val):
            card = total_card_names[i][j]
            context.list_card.verify_card_name(name_of_list, card)


@when(u'Click on the card and delete "{list_size}" and "{card_size}" and "{delete}" and "{delete_content}"')
def click_on_the_card_and_delete(context, list_size, card_size, delete, delete_content):
    context.list_card = AddAListCardPage(context.driver)
    lists_val = int(list_size)
    cards_val = int(card_size)
    for i in range(lists_val):
        name_of_list = list_names[i]
        for j in range(cards_val):
            card = total_card_names[i][j]
            context.list_card.deleting_the_card(name_of_list, card, delete, delete_content)


@when(u'Click on the list menu list size "{list_size}" and archive this list "{archive}"')
def click_on_the_list_menu_and_archive_this_list(context, list_size, archive):
    lists_val = int(list_size)
    for i in range(lists_val):
        name_of_list = list_names[i]
        context.list_card.click_list_menu_button(name_of_list)
        for row in context.table:
            header_text = row['headers']
            context.list_card.verify_list_actions_and_archive(header_text)
        context.list_card.click_on_archive_the_list_option(archive)


@then(u'Validate deleted lists list size "{list_size}" in the board')
def validate_deleted_lists_in_the_board(context, list_size):
    lists_val = int(list_size)
    for i in range(lists_val):
        name_of_list = list_names[i]
        context.list_card.verify_deleted_lists(name_of_list)


@when(u'Click on the card name "{list_size}" "{card_size}" add "{description}" and "{save}" in the card details form')
def click_on_the_card_name_and_add_description(context, list_size, card_size, description, save):
    context.list_card = AddAListCardPage(context.driver)
    lists_val = int(list_size)
    cards_val = int(card_size)
    for i in range(lists_val):
        name_of_list = list_names[i]
        for j in range(cards_val):
            card_name = total_card_names[i][j]
            fake = Faker()
            num_words_to_generate = 3
            generated_words = [fake.word() for _ in range(num_words_to_generate)]
            words_as_string = ' '.join(generated_words)
            context.list_card.click_on_card_and_add_description_to_card(name_of_list, card_name, description,
                                                                        words_as_string, save)


@when(u'Click on edit "{edit}" update the description save "{save}" list size "{list_size}" card size "{card_size}"')
def click_on_the_card_name_and_update_description(context, edit, save, list_size, card_size):
    context.list_card = AddAListCardPage(context.driver)
    lists_val = int(list_size)
    cards_val = int(card_size)
    for i in range(lists_val):
        list_name = list_names[i]
        for j in range(cards_val):
            card_name = total_card_names[i][j]
            fake = Faker()
            num_words_to_generate = 5
            generated_words = [fake.word() for _ in range(num_words_to_generate)]
            update_string = ' '.join(generated_words)
            context.list_card.click_on_card_name_and_update_description(list_name, card_name, edit, update_string, save)


@when(u'Click on the attachment of card "{list_size}" "{card_size}" and upload image')
def click_on_the_attachment_and_upload_image(context, list_size, card_size):
    context.list_card = AddAListCardPage(context.driver)
    directory_path = str(Path().absolute())
    file_path = directory_path + "\input_files\image1.png"
    lists_val = int(list_size)
    cards_val = int(card_size)
    for i in range(lists_val):
        list_name = list_names[i]
        for j in range(cards_val):
            card_name = total_card_names[i][j]
            context.list_card.click_on_attachment_upload_image(list_name, card_name, file_path)
            context.list_card.verify_success_text()
            context.list_card.click_on_close_btn()
            context.list_card.click_on_close_window()


@then(u'Validate attached image to card "{list_size}" "{card_size}" and select color and add label name')
def validate_attached_image_to_card_select_color_add_label_name(context, list_size, card_size):
    context.list_card = AddAListCardPage(context.driver)
    lists_val = int(list_size)
    cards_val = int(card_size)
    for i in range(lists_val):
        list_name = list_names[i]
        for j in range(cards_val):
            card_name = total_card_names[i][j]
            context.list_card.click_on_current_card(list_name, card_name)
            context.list_card.verify_uploaded_image_for_card()
            for row in context.table:
                color = row['color']
                total_colors.append(color)
            label_name = fake.name()
            label_names = label_name + " label"
            total_label_names.append(label_names)
            context.list_card.add_color_and_label_name(total_colors[j], label_names)
            context.list_card.click_on_close_window()


@then(u"Validate label name and colors")
def validate_label_name_and_color(context):
    context.list_card = AddAListCardPage(context.driver)
    for i in range(len(total_label_names)):
        context.list_card.verify_color_name_of_label(total_colors[i], total_label_names[i])


@when(u'Create lists "{input}" "{filename}" and validate lists')
def create_list_title_validate_list_name(context, input, filename):
    context.list_card = AddAListCardPage(context.driver)
    file_path = os.path.join(os.getcwd(), "resources")
    context.filename = filename
    context.listfilepath = os.path.join(file_path, filename)
    value = int(input)
    with open(context.listfilepath, 'w') as file:
        file.write('')
    for i in range(value):
        name = fake.name()
        with open(context.listfilepath, 'a') as file:
            file.write(f'{name}\n')
    with open(context.listfilepath, 'r') as file:
        lines = file.readlines()
        for line in lines:
            name = line.strip()
            context.list_card.enter_list_title(name)
            context.list_card.click_on_add_a_list_button()
            context.list_card.verify_added_list_name(name)


@when(u'Create cards "{list_size}" "{card_size}" "{filename}" and enter card names')
def click_on_add_a_card_button_enter_card_name(context, list_size, card_size, filename):
    context.list_card = AddAListCardPage(context.driver)
    list_val = int(list_size)
    cards_val = int(card_size)
    file_path = os.path.join(os.getcwd(), "resources")
    context.filename = filename
    context.cardfilepath = os.path.join(file_path, filename)

    with open(context.cardfilepath, 'w') as file:
        file.write('')

    for i in range(list_val * cards_val):
        card_name = fake.name()
        with open(context.cardfilepath, 'a') as file:
            file.write(f'{card_name}\n')

    with open(context.listfilepath, 'r') as file:
        list_names = file.readlines()

    for i in range(list_val):
        current_list = list_names[i].strip()
        context.list_card.click_on_add_a_card_button(current_list)

        with open(context.cardfilepath, 'r') as file:
            card_names = file.readlines()
            batch_card_names = card_names[i * cards_val: (i + 1) * cards_val]

            for card_name in batch_card_names:
                name_card = card_name.strip()
                context.list_card.enter_card_name(name_card)
                context.list_card.click_on_add_card_button()


@then(u'Validate cards "{card_size}" based on lists "{list_size}"')
def validate_created_cards_under_lists(context, card_size, list_size):
    context.list_card = AddAListCardPage(context.driver)
    cards_val = int(card_size)
    list_val = int(list_size)
    with open(context.listfilepath, 'r') as file:
        list_names = file.readlines()
    for i in range(list_val):
        current_list = list_names[i].strip()
        with open(context.cardfilepath, 'r') as file:
            card_names = file.readlines()
            batch_card_names = card_names[i * cards_val: (i + 1) * cards_val]

            for card_name in batch_card_names:
                name_card = card_name.strip()
                context.list_card.verify_card_name(current_list, name_card)


@when(u'Create lists "{list_size}" "{listfilename}" validate added lists')
def create_list_title_validate_list_name(context, list_size, listfilename):
    context.list_card = AddAListCardPage(context.driver)
    file_path = os.path.join(os.getcwd(), "resources")
    context.filename = listfilename
    context.listfilepath = os.path.join(file_path, listfilename)
    list_value = int(list_size)

    wb = Workbook()
    ws = wb.active
    ws.delete_cols(1, ws.max_column)
    ws['A1'] = 'List Names'
    for i in range(2, list_value + 2):
        name = fake.name()
        ws.cell(row=i, column=1, value=name)
    wb.save(context.listfilepath)
    names_list = []
    for row in ws.iter_rows(min_row=2, max_col=1, max_row=list_value + 1, values_only=True):
        for cell_value in row:
            if cell_value:
                names_list.append(cell_value)
    for names in names_list:
        list_name = names
        context.list_card.enter_list_title(list_name)
        context.list_card.click_on_add_a_list_button()
        context.list_card.verify_added_list_name(list_name)
